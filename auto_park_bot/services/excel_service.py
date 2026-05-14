import os
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.shift_service import get_shifts_by_period


def get_periods():
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)

    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    month_start = today.replace(day=1)

    return {
        "Сегодня": (today, today),
        "Вчера": (yesterday, yesterday),
        "Текущая неделя": (week_start, week_end),
        "Текущий месяц": (month_start, today),
    }


def style_sheet(ws):
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")

    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "ИТОГО":
                for total_cell in ws[cell.row]:
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    total_cell.border = border

    money_columns = ["E", "F", "G", "H", "I", "J"]

    for col in money_columns:
        for cell in ws[col]:
            if cell.row >= 5:
                cell.number_format = '#,##0 "сом"'

    widths = {
        "A": 12,
        "B": 24,
        "C": 16,
        "D": 12,
        "E": 16,
        "F": 18,
        "G": 18,
        "H": 16,
        "I": 16,
        "J": 16,
        "K": 24,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"


async def fill_sheet(ws, title: str, start_date: str, end_date: str):
    shifts = await get_shifts_by_period(start_date, end_date)

    ws.merge_cells("A1:K1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Период: {start_date} — {end_date}"
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "ID смены",
        "Водитель",
        "Машина",
        "Заказы",
        "Оборот",
        "Расход зарядки",
        "Комиссия Яндекса",
        "Зарплата 30%",
        "Все расходы",
        "Прибыль",
        "Дата закрытия"
    ]

    ws.append([])
    ws.append(headers)

    total_orders = 0
    total_turnover = 0
    total_charging = 0
    total_yandex = 0
    total_salary = 0
    total_expenses = 0
    total_profit = 0

    for shift in shifts:
        shift_id = shift[0]
        driver_name = shift[1]
        plate_number = shift[2]
        orders_count = shift[3] or 0
        turnover = shift[4] or 0
        charging = shift[5] or 0
        yandex = shift[6] or 0
        salary = shift[7] or 0
        end_time = shift[8]

        expenses = charging + yandex + salary
        profit = turnover - expenses

        total_orders += orders_count
        total_turnover += turnover
        total_charging += charging
        total_yandex += yandex
        total_salary += salary
        total_expenses += expenses
        total_profit += profit

        ws.append([
            shift_id,
            driver_name,
            plate_number,
            orders_count,
            turnover,
            charging,
            yandex,
            salary,
            expenses,
            profit,
            end_time
        ])

    ws.append([])
    ws.append([
        "ИТОГО",
        "",
        "",
        total_orders,
        total_turnover,
        total_charging,
        total_yandex,
        total_salary,
        total_expenses,
        total_profit,
        ""
    ])

    style_sheet(ws)


async def create_full_excel_report():
    os.makedirs("exports", exist_ok=True)

    periods = get_periods()

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, dates in periods.items():
        start_date = dates[0].strftime("%Y-%m-%d")
        end_date = dates[1].strftime("%Y-%m-%d")

        ws = wb.create_sheet(title=sheet_name)
        await fill_sheet(ws, sheet_name, start_date, end_date)

    month_name = datetime.now().strftime("%Y_%m")
    filename = f"report_{month_name}.xlsx"
    filepath = os.path.join("exports", filename)

    wb.save(filepath)

    return filepath