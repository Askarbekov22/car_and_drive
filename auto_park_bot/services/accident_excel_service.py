import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from services.accident_service import get_accidents_for_excel


PAYMENT_NAMES = {
    "driver": "Водитель",
    "company": "Компания",
    "insurance": "КАСКО",
    "mixed": "Смешанная оплата",
}

STATUS_NAMES = {
    "open": "в процессе",
    "closed": "выплачен",
}


def _safe_number(value):
    if value is None:
        return None

    try:
        number = float(value)
    except (TypeError, ValueError):
        return None

    if number == 0:
        return None

    return number


def _style_header(ws, row_number=1):
    fill = PatternFill("solid", fgColor="D9EAD3")
    font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[row_number]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_table(ws):
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"


def _set_widths(ws, widths):
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def generate_accident_excel_path() -> str:
    exports_dir = "exports"
    os.makedirs(exports_dir, exist_ok=True)

    filename = f"dtp_report_{datetime.now().strftime('%Y_%m_%d_%H_%M_%S')}.xlsx"
    return os.path.join(exports_dir, filename)


async def create_accident_excel() -> str:
    accidents = await get_accidents_for_excel()

    wb = Workbook()

    ws_main = wb.active
    ws_main.title = "Учет ДТП"

    ws_repair = wb.create_sheet("Ремонт")
    ws_debtors = wb.create_sheet("Должники")

    _fill_main_sheet(ws_main, accidents)
    _fill_repair_sheet(ws_repair, accidents)
    _fill_debtors_sheet(ws_debtors, accidents)

    path = generate_accident_excel_path()
    wb.save(path)

    return path


def _fill_main_sheet(ws, accidents):
    headers = [
        "Номер №",
        "Дата",
        "Машина",
        "Водитель",
        "Место",
        "Тип Дтп",
        "Виновен",
        "Факт Ремонт",
        "Оценка",
        "Кто платит",
        "Статус",
        "Простой (смены)",
        "Потери",
        "Разница ",
    ]

    ws.append(headers)

    for index, accident in enumerate(accidents, start=2):
        row_number = index
        payment_type = accident.get("payment_type")
        status = accident.get("status")

        ws.append([
            accident.get("id"),
            accident.get("accident_date"),
            accident.get("plate_number"),
            accident.get("full_name"),
            accident.get("accident_place"),
            accident.get("accident_type"),
            accident.get("guilty_party"),
            _safe_number(accident.get("repair_fact")),
            _safe_number(accident.get("estimate_amount")),
            PAYMENT_NAMES.get(payment_type, payment_type),
            STATUS_NAMES.get(status, status),
            _safe_number(accident.get("downtime_shifts")),
            _safe_number(accident.get("losses_amount")),
            f"=I{row_number}-H{row_number}-M{row_number}",
        ])

    _style_header(ws)
    _style_table(ws)

    _set_widths(ws, [
        8.75,
        12,
        10,
        24,
        20,
        37,
        13,
        13,
        13,
        24,
        15,
        16,
        13,
        13,
    ])

    for row in range(2, ws.max_row + 1):
        ws[f"B{row}"].number_format = "yyyy-mm-dd"
        ws[f"H{row}"].number_format = '#,##0'
        ws[f"I{row}"].number_format = '#,##0'
        ws[f"M{row}"].number_format = '#,##0'
        ws[f"N{row}"].number_format = '#,##0'


def _fill_repair_sheet(ws, accidents):
    headers = [
        "ДТП №",
        "Машина",
        "Водитель",
        "Дата",
        "Тип ДТП",
        "Факт ремонт",
        "Оценка",
        "Разница",
    ]

    ws.append(headers)

    for accident in accidents:
        ws.append([
            accident.get("id"),
            accident.get("plate_number"),
            accident.get("full_name"),
            accident.get("accident_date"),
            accident.get("accident_type"),
            _safe_number(accident.get("repair_fact")),
            _safe_number(accident.get("estimate_amount")),
            _safe_number(accident.get("difference_amount")),
        ])

    _style_header(ws)
    _style_table(ws)

    _set_widths(ws, [
        10,
        12,
        24,
        12,
        35,
        15,
        15,
        15,
    ])

    for row in range(2, ws.max_row + 1):
        ws[f"D{row}"].number_format = "yyyy-mm-dd"
        ws[f"F{row}"].number_format = '#,##0'
        ws[f"G{row}"].number_format = '#,##0'
        ws[f"H{row}"].number_format = '#,##0'


def _fill_debtors_sheet(ws, accidents):
    headers = [
        "",
        "Машина",
        "ФИО",
        "Долг",
        "Отдал",
        "Остаток",
    ]

    ws.append(headers)

    row_index = 1

    for accident in accidents:
        debt_amount = accident.get("debt_amount") or 0

        if debt_amount <= 0:
            continue

        row_index += 1

        ws.append([
            row_index - 1,
            accident.get("plate_number"),
            accident.get("full_name"),
            _safe_number(accident.get("estimate_amount")),
            _safe_number(accident.get("paid_amount")),
            f"=D{row_index}-E{row_index}",
        ])

    _style_header(ws)
    _style_table(ws)

    _set_widths(ws, [
        5,
        12,
        22,
        13,
        13,
        13,
    ])

    for row in range(2, ws.max_row + 1):
        ws[f"D{row}"].number_format = '#,##0'
        ws[f"E{row}"].number_format = '#,##0'
        ws[f"F{row}"].number_format = '#,##0'